#!/usr/bin/env python3
"""enrich.py — Filter and enrich the ECPC repository spreadsheet.

1. Removes known false-positive members.
2. Applies an inclusive ECPC-relevance filter (child / caregiver / conflict /
   peacebuilding / ECD topics).
3. For OPEN-ACCESS papers: downloads PDFs, reads methods sections, and
   extracts country, sample size, age, demographics, effect sizes, key findings,
   and study design.
4. For NON-OPEN-ACCESS papers: keeps them with available metadata and adds
   access_type = "Subscription Required" plus manual_lookup = "Y" so they
   can be found manually.
5. Resolves "unclear" design labels using a comprehensive rule set.
6. Saves a clean, NUL-free CSV.

Usage:
    python enrich.py ecpc_2023_2026.xlsx ecpc_enriched.csv
"""

import csv
import io
import re
import sys
import time
from pathlib import Path
from typing import Optional

import openpyxl
import requests

try:
    from pypdf import PdfReader
    HAS_PYPDF = True
except ImportError:
    HAS_PYPDF = False

# ── Known false-positive members (wrong OpenAlex match, no ORCID) ──────────

FALSE_POSITIVE_MEMBERS: set[tuple[str, str]] = {
    ("benjamin", "dave"),
    ("rolland", "eduardo"),
    ("johnson", "isaac"),
    ("perez", "elizabeth"),
}

# ── ECPC relevance keywords ────────────────────────────────────────────────

_CHILD_TERMS = [
    "child", "children", "childhood", "infant", "infancy", "toddler",
    "preschool", "pre-school", "kindergarten", "early childhood",
    "newborn", "neonate", "neonatal", "baby", "babies",
    "young child", "pediatric", "perinatal", "birth cohort",
    "school-age", "school age", "youth", "adolescent",
]
_CAREGIVER_TERMS = [
    "maternal", "mother", "father", "parent", "caregiver", "caregiver",
    "pregnant", "postnatal", "antenatal", "prenatal", "family",
    "dyad", "two-generation",
]
_CONFLICT_HUMANITARIAN = [
    "conflict", "war", "post-conflict", "post conflict", "armed conflict",
    "humanitarian", "refugee", "displaced", "displacement", "asylum seeker",
    "asylum-seeker", "fragile state", "fragile setting", "emergency setting",
    "crisis", "genocide", "mass atrocity", "political violence",
    "peace", "peacebuilding", "peace building", "reconciliation",
    "reconstruction", "transitional justice", "psychosocial support",
    # Specific conflict zones commonly studied by ECPC members
    "gaza", "west bank", "palestine", "palestinian",
    "ukraine", "syria", "syrian", "rwanda", "sierra leone", "liberia",
    "south sudan", "rohingya", "myanmar", "lebanon", "iraq",
    "afghanistan", "burundi", "congo", "somalia", "yemen", "mali",
    "central african republic",
]
_DOMAIN_TERMS = [
    "early childhood development", "ecd", "early childhood care",
    "mental health", "psychosocial", "wellbeing", "well-being",
    "development", "developmental",
    "nurturing care", "responsive care",
    "violence against children", "child abuse", "maltreatment",
    "corporal punishment", "physical punishment", "child protection",
    "trauma", "adversity", "resilience", "attachment",
    "nutrition", "stunting", "wasting", "anthropometric",
    "education", "learning", "literacy", "school readiness",
    "social-emotional", "social emotional", "executive function",
    "depression", "anxiety", "ptsd", "post-traumatic",
    "intervention", "programme evaluation", "program evaluation",
    "poverty", "cash transfer", "economic inclusion",
    "parenting", "parenting program", "home visiting",
    "disability", "special needs",
]

# Hard exclusions — clearly non-ECPC fields
_EXCLUSION_PATTERNS = [
    r"\bastrophysic|\bgalax[yi]|\bstellar\b|\bblack hole\b|\bastronomic|\btelescope\b",
    r"\batmospheric\b.*\bcloud|\bcloud droplet|\baerosol\b|\bprecipitation\b.*\bradar|\bmicrophysics\b|\bice crystal\b",
    r"\bsnakebite|\bsnake venom|\benvenomation|\bantivenom\b|\bviper\b.*\bvenom",
    r"\bparasit|\bnematode\b|\bhelminth\b|\bcopepoda\b|\btrypanosomatid\b|\bbat\b.*\bparasit",
    r"\bancolog|\btumou?r\b.*\bcancer|\bmalignancy\b|\bmetastasis\b|\bchemotherapy\b",
    r"\bperitonectomy\b|\blymphadenectomy\b|\bsurgical technique\b|\bendoscop",
    r"\bhigh.energy physics\b|\bparticle physics\b|\bquark\b|\bboson\b",
    r"\bwine\b.*\bferment|\bbeer\b.*\bferment|\bbrewery\b",
    r"\bgeolog[yi]|\bearthquake\b|\bseismic\b|\btectonic\b",
]


def _any_in(text: str, terms: list) -> bool:
    return any(t in text for t in terms)


def is_ecpc_relevant(title: str, abstract: str) -> bool:
    """Return True if paper is within ECPC scope."""
    text = (title + " " + (abstract or "")).lower()

    for pat in _EXCLUSION_PATTERNS:
        if re.search(pat, text, re.IGNORECASE):
            return False

    has_child = _any_in(text, _CHILD_TERMS)
    has_caregiver = _any_in(text, _CAREGIVER_TERMS)
    has_conflict = _any_in(text, _CONFLICT_HUMANITARIAN)
    has_domain = _any_in(text, _DOMAIN_TERMS)

    if has_child and has_domain:
        return True
    if has_child and has_caregiver:
        return True
    if has_caregiver and has_domain and (has_conflict or "intervention" in text):
        return True
    if has_conflict and (has_child or has_caregiver or has_domain):
        return True
    if "peacebuilding" in text or "peace building" in text or "peace-building" in text:
        return True
    if "early childhood" in text:
        return True
    if "psychosocial support" in text:
        return True
    return False


# ── Comprehensive design detection ────────────────────────────────────────

# (pattern, design_label, randomized_label)
# ORDER MATTERS — first match wins.
# Priority: reviews → qualitative/mixed → interventions → observational
_DESIGN_RULES: list[tuple[str, str, str]] = [
    # ── Reviews first (they cite RCTs/longitudinal — must not be mis-labelled) ──
    (r"meta.analy", "meta-analysis", "not applicable"),
    (r"systematic review|scoping review|cochrane review", "systematic review", "not applicable"),
    (r"rapid review|integrative review|realist review|"
     r"\bnarrative review\b|\bliterature review\b|"
     r"\breview of.{1,30}literature\b|\bstate of.{1,20}evidence\b|"
     r"\breview of evidence\b|\bscoping literature\b", "review", "not applicable"),
    # ── Qualitative & mixed before intervention/longitudinal ──
    (r"mixed.method|mixed method|"
     r"qualitative and quantitative|quantitative and qualitative|"
     r"convergent design|explanatory sequential|exploratory sequential", "mixed-methods", "no"),
    (r"qualitative (study|research|method|data|approach|inquiry|component)|"
     r"thematic analysis|grounded theory|phenomenolog|ethnograph|"
     r"discourse analysis|narrative inquiry|interpretive (study|research)|"
     r"in.depth interview|semi.structured interview|"
     r"focus group|photovoice|"
     r"participatory action research|"
     r"we conducted (interviews|focus groups|in.depth)|"
     r"we interviewed|we observed", "qualitative", "not applicable"),
    # ── Randomised interventions ──
    (r"cluster.randomi[sz]ed|cluster rct|c?rct\b.*cluster|stepped.wedge", "cluster RCT", "yes"),
    (r"randomized controlled trial|randomised controlled trial|\brct\b|\(rct\)|"
     r"randomly assigned|random assignment|random allocation|"
     r"double.blind.*trial|single.blind.*trial|placebo.controlled trial|"
     r"we randomized|were randomized|were randomly allocated", "RCT", "yes"),
    # ── Quasi-experimental ──
    (r"quasi.experimental|difference.in.difference|regression discontinuity|"
     r"interrupted time series|natural experiment|"
     r"propensity score match|instrumental variable|"
     r"synthetic control", "quasi-experimental", "no"),
    # ── Secondary / archival (before longitudinal) ──
    (r"secondary analysis|secondary data analysis|"
     r"archival data|administrative data|existing data|"
     r"data from.{1,30}(randomized|rct|cohort|survey|trial)|"
     r"previously collected data|retrospective (cohort|data|analysis|review)", "secondary analysis", "no"),
    # ── Feasibility / pilot ──
    (r"feasibility (study|trial|test)|pilot (study|trial|rct)|"
     r"feasibility and (acceptability|pilot)|proof of concept", "feasibility/pilot", "unclear"),
    # ── Program / process evaluation ──
    (r"process evaluation|formative evaluation|"
     r"implementation (study|evaluation|science|research)|"
     r"program (evaluation|assessment)|programme (evaluation|assessment)|"
     r"quality improvement", "program evaluation", "no"),
    # ── Case study ──
    (r"\bcase study\b|\bcase studies\b|\bcase report\b|\bcase series\b|\bcase.control\b", "case study/report", "not applicable"),
    # ── Longitudinal / prospective ──
    (r"birth cohort|prospective cohort|prospective longitudinal|"
     r"longitudinal cohort|longitudinal study|longitudinal follow.?up|"
     r"panel study|panel data|repeated measures|multiple waves|"
     r"followed.{1,20}over|tracked.{1,30}over time|data at.{1,15}time point", "longitudinal cohort", "no"),
    # ── Cross-sectional ──
    (r"cross.sectional|prevalence study|nationally representative survey|"
     r"cross.national survey|household survey|community survey|"
     r"population.based survey|descriptive survey", "cross-sectional", "no"),
    # ── Observational / descriptive ──
    (r"\bobservational (study|design|research)\b|\bdescriptive (study|research)\b|"
     r"\bexploratory (study|research)\b|\bnaturalistic (study|research)\b", "observational", "no"),
    # ── Commentary / conceptual (non-empirical) ──
    (r"\bthis commentary\b|\bthis editorial\b|\bthis perspective\b|"
     r"\bopinion (piece|paper)\b|\bviewpoint\b|\bcall to action\b|"
     r"\bpolicy brief\b|\bpractitioner note\b|"
     r"\bconceptual (framework|paper|model)\b|\btheoretical (framework|paper)\b", "commentary/conceptual", "not applicable"),
    # ── Weak signals — last resort ──
    (r"\blongitudinal\b|\bprospective\b.{1,30}\bstudy\b|\bfollow.?up\b", "longitudinal cohort", "no"),
    (r"\bsurvey\b.{1,30}(children|families|caregivers|parents|schools|youth|adolescent)", "cross-sectional", "no"),
    (r"\bcohort\b", "longitudinal cohort", "no"),
]

_COMPILED_RULES = [(re.compile(pat, re.IGNORECASE), label, rand) for pat, label, rand in _DESIGN_RULES]


def detect_design_comprehensive(title: str, abstract: str, methods_text: str = "") -> tuple[str, str]:
    """Return (design, randomized) using ordered rule matching."""
    # Combine all available text; weight methods section first
    combined = (methods_text + " " + title + " " + (abstract or "")).lower()
    for compiled, label, rand in _COMPILED_RULES:
        if compiled.search(combined):
            return label, rand
    return "unclear", "unclear"


def extract_methods_section(text: str) -> str:
    """Extract text following a Methods/Design section header."""
    pat = re.compile(
        r'(?:^|\n)\s*(?:\d[\.\d]*\s+)?'
        r'(?:Methods?|Study Design|Methodology|Design|Materials? and Methods?|'
        r'Methods? and Materials?|Participants? and Procedures?)\s*\n',
        re.IGNORECASE | re.MULTILINE,
    )
    m = pat.search(text)
    if m:
        chunk = text[m.end(): m.end() + 1500]
        # Stop at next section header
        stop = re.search(r'\n\s*(?:\d[\.\d]*\s+)?[A-Z][A-Za-z ]{3,30}\s*\n', chunk)
        if stop:
            chunk = chunk[: stop.start()]
        return chunk
    return ""


# ── Extraction helpers ─────────────────────────────────────────────────────

_COUNTRY_LIST = [
    "Afghanistan", "Albania", "Algeria", "Angola", "Argentina", "Armenia",
    "Australia", "Austria", "Azerbaijan", "Bangladesh", "Belgium", "Belize",
    "Bolivia", "Bosnia", "Botswana", "Brazil", "Burkina Faso", "Burundi",
    "Cambodia", "Cameroon", "Canada", "Central African Republic",
    "Chad", "Chile", "China", "Colombia", "Congo", "Costa Rica",
    "Côte d'Ivoire", "Cote d'Ivoire", "Croatia", "Cuba",
    "Democratic Republic", "DR Congo", "DRC",
    "Ecuador", "Egypt", "El Salvador", "Ethiopia",
    "Gaza", "Georgia", "Ghana", "Guatemala", "Guinea",
    "Guinea-Bissau", "Haiti", "Honduras", "India", "Indonesia",
    "Iran", "Iraq", "Israel", "Jamaica", "Jordan", "Kazakhstan",
    "Kenya", "Kosovo", "Laos", "Lebanon", "Lesotho", "Liberia",
    "Libya", "Malawi", "Mali", "Mexico", "Moldova", "Mongolia",
    "Morocco", "Mozambique", "Myanmar", "Namibia", "Nepal",
    "Nicaragua", "Niger", "Nigeria", "Norway", "Pakistan",
    "Palestine", "Palestinian", "Panama", "Papua New Guinea", "Peru",
    "Philippines", "Poland", "Portugal", "Rwanda", "Rohingya",
    "Senegal", "Serbia", "Sierra Leone", "Somalia", "South Africa",
    "South Sudan", "Sri Lanka", "Sudan", "Syria", "Tanzania",
    "Thailand", "Timor-Leste", "Togo", "Tunisia", "Turkey",
    "Uganda", "Ukraine", "United Kingdom", "United States",
    "Uruguay", "Venezuela", "Vietnam", "West Bank", "Yemen",
    "Zambia", "Zimbabwe",
    # Regional / descriptive
    "Sub-Saharan Africa", "sub-Saharan", "Latin America",
    "Middle East", "Southeast Asia", "South Asia",
    "East Africa", "West Africa", "Central America",
    "Northern Ireland", "LMIC", "low- and middle-income",
    "low-income countr", "middle-income countr",
    "conflict-affected", "conflict affected", "war-affected",
]
_COUNTRY_RE = re.compile(
    r"\b(" + "|".join(re.escape(c) for c in sorted(_COUNTRY_LIST, key=len, reverse=True)) + r")\b",
    re.IGNORECASE,
)

_SAMPLE_PATS = [
    re.compile(r'\bn\s*=\s*([\d,]+)', re.IGNORECASE),
    re.compile(r'\bN\s*=\s*([\d,]+)'),
    re.compile(r'([\d,]+)\s+(?:participants?|children|infants?|families|caregivers?|mothers?|fathers?|dyads?|households?|adolescents?|youth|women|men)\b', re.IGNORECASE),
    re.compile(r'sample(?:\s+size)?\s+(?:of\s+)?([\d,]+)', re.IGNORECASE),
    re.compile(r'(?:enrolled|recruited|included|randomized|randomised)\s+([\d,]+)', re.IGNORECASE),
    re.compile(r'total\s+(?:of\s+)?([\d,]+)\s+(?:participants?|children|families)', re.IGNORECASE),
]
_AGE_PATS = [
    re.compile(r'aged?\s+([\d.]+)\s*(?:to|[-–])\s*([\d.]+)\s*years?', re.IGNORECASE),
    re.compile(r'ages?\s+([\d.]+)\s*(?:to|[-–])\s*([\d.]+)\s*years?', re.IGNORECASE),
    re.compile(r'age\s+range\s+(?:of\s+)?([\d.]+)\s*(?:to|[-–])\s*([\d.]+)', re.IGNORECASE),
    re.compile(r'mean\s+age\s+(?:was\s+|of\s+)?([\d.]+)', re.IGNORECASE),
    re.compile(r'([\d.]+)\s*(?:to|[-–])\s*([\d.]+)\s*months?\s+(?:of age|old)', re.IGNORECASE),
    re.compile(r'([\d.]+)\s*(?:to|[-–])\s*([\d.]+)[- ]year[- ]old', re.IGNORECASE),
    re.compile(r'birth\s+to\s+(\d+)\s*years?', re.IGNORECASE),
]
_DEMO_FLAGS = [
    (re.compile(r'\b(refugee[s]?|displaced persons?|internally displaced|IDP[s]?|asylum seeker[s]?)\b', re.IGNORECASE), "refugees/IDPs"),
    (re.compile(r'\b(low.income|impoverished|poor household[s]?|living in poverty)\b', re.IGNORECASE), "low-income"),
    (re.compile(r'\b(urban\b|rural\b|peri.urban)\b', re.IGNORECASE), None),
    (re.compile(r'\b(girl[s]?|boy[s]?|gender)\b', re.IGNORECASE), None),
    (re.compile(r'\b(ethnic|indigenous|minority|marginalized|marginalised|underserved)\b', re.IGNORECASE), None),
    (re.compile(r'\b(orphan[s]?|unaccompanied|separated children?)\b', re.IGNORECASE), None),
    (re.compile(r'\b(caregiver[s]?|mother[s]?|father[s]?|parent[s]?)\b', re.IGNORECASE), None),
    (re.compile(r'\b(single.parent|single parent|lone parent)\b', re.IGNORECASE), "single-parent"),
    (re.compile(r'\b(disability|disabled|special needs|developmental delay)\b', re.IGNORECASE), None),
    (re.compile(r'\b(conflict.affected|war.affected|post.conflict)\b', re.IGNORECASE), "conflict-affected"),
]
_ES_PATS = [
    re.compile(r"Cohen[''s]*\s*d\s*=\s*([-+]?\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"Hedges[''s]*\s*g\s*=\s*([-+]?\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"\bd\s*=\s*([-+]?\d+\.\d+)"),
    re.compile(r"\bg\s*=\s*([-+]?\d+\.\d+)"),
    re.compile(r"\br\s*=\s*([-+]?\d+\.\d+)"),
    re.compile(r"\bβ\s*=\s*([-+]?\d+\.\d+)"),
    re.compile(r"\bOR\s*=\s*(\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"\bRR\s*=\s*(\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"\bHR\s*=\s*(\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"\bSMD\s*=\s*([-+]?\d+\.?\d*)", re.IGNORECASE),
    re.compile(r"\bp\s*[<>=]\s*0?\.\d+"),
]
_FINDING_RE = re.compile(
    r'(?:we\s+found|results?\s+(?:showed?|indicated?|demonstrated?|revealed?|suggest)|'
    r'findings?\s+(?:showed?|indicated?|suggested?|revealed?)|'
    r'(?:significant(?:ly)?|substantially)\s+(?:higher|lower|improved?|reduced?|increased?|greater|less)|'
    r'(?:children|participants?|caregivers?|mothers?|families)\s+(?:showed?|demonstrated?|reported?|had\s+significant)|'
    r'the\s+(?:intervention|program|programme|treatment)\s+(?:significantly|substantially|led|resulted)|'
    r'there\s+was\s+a\s+significant)',
    re.IGNORECASE,
)


def extract_country(text: str) -> str:
    found = dict.fromkeys(m.group(1).title() for m in _COUNTRY_RE.finditer(text[:5000]))
    return "; ".join(list(found)[:6])


def extract_sample_size(text: str) -> str:
    cands: list[int] = []
    for pat in _SAMPLE_PATS:
        for m in pat.finditer(text[:6000]):
            raw = m.group(1).replace(",", "")
            try:
                n = int(raw)
                if 5 <= n <= 2_000_000:
                    cands.append(n)
            except ValueError:
                pass
    if not cands:
        return ""
    return str(max(set(cands), key=cands.count))


def extract_age(text: str) -> str:
    results: list[str] = []
    for pat in _AGE_PATS:
        m = pat.search(text[:6000])
        if m:
            g = m.groups()
            results.append(f"{g[0]}–{g[1]}" if len(g) == 2 else g[0])
            if len(results) >= 2:
                break
    return "; ".join(dict.fromkeys(results))


def extract_demographics(text: str) -> str:
    found: list[str] = []
    for pat, label in _DEMO_FLAGS:
        m = pat.search(text[:6000])
        if m:
            found.append(label if label else m.group(0).lower())
    return "; ".join(list(dict.fromkeys(found))[:7])


def extract_effect_sizes(text: str) -> str:
    hits: list[str] = []
    for pat in _ES_PATS:
        for m in pat.finditer(text[:8000]):
            hits.append(m.group(0))
            if len(hits) >= 8:
                break
        if len(hits) >= 8:
            break
    return "; ".join(dict.fromkeys(hits))


def extract_key_findings(abstract: str) -> tuple[str, str]:
    if not abstract:
        return "", ""
    sentences = re.split(r'(?<=[.!?])\s+', abstract.strip())
    findings: list[str] = []
    for sent in sentences:
        if _FINDING_RE.search(sent) and len(sent) > 40:
            findings.append(sent.strip())
            if len(findings) == 2:
                break
    if not findings:
        tail = [s.strip() for s in sentences[-3:] if len(s.strip()) > 40]
        findings = tail[:2]
    return (findings[0] if findings else ""), (findings[1] if len(findings) > 1 else "")


# ── PDF download ───────────────────────────────────────────────────────────

PDF_CACHE = Path("pdf_cache")
_HEADERS = {"User-Agent": "ECPC-Pipeline/1.0 (mailto:Sascha.Hein@fu-berlin.de)"}


def fetch_pdf_text(url: str, doi: str) -> str:
    if not HAS_PYPDF:
        return ""
    PDF_CACHE.mkdir(exist_ok=True)
    safe = (doi.replace("/", "_").replace(":", "_")[:80] if doi else url[-50:])
    safe = re.sub(r'[^\w._-]', '_', safe)
    dest = PDF_CACHE / f"{safe}.pdf"

    if not dest.exists():
        try:
            resp = requests.get(url, headers=_HEADERS, timeout=25, allow_redirects=True)
            if resp.status_code == 200 and b"%PDF" in resp.content[:1024]:
                dest.write_bytes(resp.content)
            else:
                return ""
        except Exception:
            return ""

    try:
        reader = PdfReader(str(dest))
        text = ""
        for page in reader.pages[:10]:
            text += (page.extract_text() or "") + "\n"
            if len(text) > 10000:
                break
        return text[:10000].replace("\x00", "")
    except Exception:
        return ""


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) < 3:
        print("Usage: python enrich.py INPUT.xlsx OUTPUT.csv")
        sys.exit(1)

    input_path, output_path = Path(sys.argv[1]), Path(sys.argv[2])
    print(f"Loading {input_path} ...")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    def col(row, name: str) -> str:
        try:
            return str(row[headers.index(name)] or "")
        except (ValueError, IndexError):
            return ""

    total = len(raw_rows)
    n_fp = n_irrel = 0
    kept: list[dict] = []

    for row in raw_rows:
        last  = col(row, "Ecpc Last").lower().strip()
        first = col(row, "Ecpc First").lower().strip()
        title    = col(row, "Title")
        abstract = col(row, "Abstract")
        oa_url   = col(row, "Oa Url")
        doi      = col(row, "Doi")

        # ── 1. Drop false-positive members
        if (last, first) in FALSE_POSITIVE_MEMBERS:
            n_fp += 1
            continue

        # ── 2. Relevance filter
        if not is_ecpc_relevant(title, abstract):
            n_irrel += 1
            continue

        # ── 3. Build record from existing columns
        rec: dict = {}
        for h, v in zip(headers, row):
            rec[h] = str(v) if v is not None else ""

        # ── 4. Access type flag
        is_oa = bool(oa_url.strip())
        rec["Access Type"]           = "Open Access" if is_oa else "Subscription Required"
        rec["Manual Lookup Needed"]  = "" if is_oa else "Y"

        # ── 5. Baseline extraction from abstract alone
        rec["Country"]        = extract_country(abstract)        or rec.get("Country", "")
        rec["Sample Size"]    = extract_sample_size(abstract)    or rec.get("Sample Size", "")
        rec["Age Breakdowns"] = extract_age(abstract)            or rec.get("Age Breakdowns", "")
        rec["Demographics"]   = extract_demographics(abstract)   or rec.get("Demographics", "")
        rec["Effect Sizes"]   = extract_effect_sizes(abstract)   or rec.get("Effect Sizes", "")
        kf1, kf2 = extract_key_findings(abstract)
        rec["Key Finding 1"]  = kf1 or rec.get("Key Finding 1", "")
        rec["Key Finding 2"]  = kf2 or rec.get("Key Finding 2", "")

        # Improve design from abstract
        d_abs, r_abs = detect_design_comprehensive(title, abstract)
        if d_abs != "unclear":
            rec["Design"], rec["Randomized"] = d_abs, r_abs

        # ── 6. For OA papers: download PDF and enrich from full text
        if is_oa:
            print(f"  [PDF] {title[:65]}...")
            pdf_text = fetch_pdf_text(oa_url, doi)
            if pdf_text:
                methods_text = extract_methods_section(pdf_text)
                full = abstract + "\n\n" + methods_text + "\n\n" + pdf_text

                country = extract_country(full)
                if country:
                    rec["Country"] = country
                n = extract_sample_size(full)
                if n:
                    rec["Sample Size"] = n
                age = extract_age(full)
                if age:
                    rec["Age Breakdowns"] = age
                demo = extract_demographics(full)
                if demo:
                    rec["Demographics"] = demo
                es = extract_effect_sizes(full)
                if es:
                    rec["Effect Sizes"] = es
                kf1_f, kf2_f = extract_key_findings(full)
                if kf1_f:
                    rec["Key Finding 1"] = kf1_f
                if kf2_f:
                    rec["Key Finding 2"] = kf2_f

                # Re-run design detection with methods section having priority
                d_pdf, r_pdf = detect_design_comprehensive(title, abstract, methods_text)
                if d_pdf != "unclear":
                    rec["Design"], rec["Randomized"] = d_pdf, r_pdf

            time.sleep(0.2)

        kept.append(rec)

    # ── 7. Write CSV (NUL-safe) ────────────────────────────────────────────
    extra_cols = [
        "Country", "Sample Size", "Age Breakdowns", "Demographics",
        "Effect Sizes", "Key Finding 1", "Key Finding 2",
        "Access Type", "Manual Lookup Needed",
    ]
    all_keys = list(headers) + [k for k in extra_cols if k not in headers]

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=all_keys, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(kept)
    csv_bytes = buf.getvalue().replace("\x00", "").encode("utf-8")
    output_path.write_bytes(csv_bytes)

    # ── 8. Summary ────────────────────────────────────────────────────────
    from collections import Counter
    design_counts = Counter(r.get("Design", "") for r in kept)
    oa_count  = sum(1 for r in kept if r.get("Access Type") == "Open Access")
    sub_count = sum(1 for r in kept if r.get("Access Type") == "Subscription Required")

    print(f"\n{'─'*60}")
    print(f"Input rows           : {total}")
    print(f"Removed (false+)     : {n_fp}")
    print(f"Removed (irrelevant) : {n_irrel}")
    print(f"Kept                 : {len(kept)}")
    print(f"  Open Access        : {oa_count}")
    print(f"  Subscription only  : {sub_count}  ← flagged Manual Lookup Needed = Y")
    print(f"\nDesign breakdown:")
    for d, n in sorted(design_counts.items(), key=lambda x: -x[1]):
        print(f"  {(d or 'blank'):<30} {n}")
    print(f"\nSaved → {output_path}")


if __name__ == "__main__":
    main()
