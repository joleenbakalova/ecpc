#!/usr/bin/env python3
"""ECPC Research Repository Pipeline

Fetches publications by ECPC members from OpenAlex (Aug 1 2023 – today),
downloads open-access PDFs, and uses Claude to extract structured metadata
matching the All-Table 1.csv schema. Outputs a styled Excel workbook.

Usage:
    python pipeline.py                        # full run, all members
    python pipeline.py --dry-run              # fetch paper list only, no Claude
    python pipeline.py --member Connolly      # single member by last name
    python pipeline.py --output my_output.xlsx
"""

import argparse
import base64
import csv
import datetime
import json
import os
import sys
import time
from pathlib import Path
from typing import Optional

import anthropic
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Load .env file from the script's directory if present (ANTHROPIC_API_KEY, etc.)
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    for _line in _env_path.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _, _v = _line.partition("=")
            os.environ.setdefault(_k.strip(), _v.strip())

# ── Configuration ──────────────────────────────────────────────────────────────

DATE_FROM = "2023-08-01"
DATE_TO = datetime.date.today().isoformat()  # always today
OPENALEX_EMAIL = "Sascha.Hein@fu-berlin.de"
PDF_CACHE_DIR = Path("pdf_cache")
MODEL = "claude-opus-4-8"

# ── Member loading ─────────────────────────────────────────────────────────────


def load_members(members_csv: str, orcid_csv: Optional[str] = None) -> list[dict]:
    members = []
    with open(members_csv, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            members.append(dict(row))

    if orcid_csv and Path(orcid_csv).exists():
        orcid_map: dict[tuple, str] = {}
        with open(orcid_csv, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                key = (row["last"].strip().lower(), row["first"].strip().lower())
                orcid_map[key] = row.get("orcid", "").strip()
        for m in members:
            key = (m["last"].strip().lower(), m["first"].strip().lower())
            m["orcid"] = orcid_map.get(key, "")
    else:
        for m in members:
            m.setdefault("orcid", "")

    return members


# ── OpenAlex helpers ───────────────────────────────────────────────────────────


def find_openalex_author_id(last: str, first: str, affiliation: str) -> Optional[str]:
    """Search OpenAlex for an author; prefer affiliation match, fall back to top result."""
    url = "https://api.openalex.org/authors"
    params = {
        "search": f"{first} {last}",
        "mailto": OPENALEX_EMAIL,
        "per-page": 5,
    }
    try:
        resp = requests.get(url, params=params, timeout=15)
        resp.raise_for_status()
        results = resp.json().get("results", [])
    except Exception as e:
        print(f"  [WARN] OpenAlex author search failed: {e}")
        return None

    if not results:
        return None

    # Prefer a result whose last-known institution overlaps with the known affiliation
    aff_words = {w for w in affiliation.lower().split() if len(w) > 4}
    for author in results:
        inst = (author.get("last_known_institution") or {}).get("display_name", "")
        if aff_words and any(w in inst.lower() for w in aff_words):
            return author["id"]

    return results[0]["id"]


def reconstruct_abstract(inverted_index: Optional[dict]) -> str:
    if not inverted_index:
        return ""
    positions: list[tuple[int, str]] = []
    for word, pos_list in inverted_index.items():
        for pos in pos_list:
            positions.append((pos, word))
    positions.sort()
    return " ".join(word for _, word in positions)


def get_oa_pdf_url(work: dict) -> Optional[str]:
    oa = work.get("open_access", {})
    if oa.get("oa_url"):
        return oa["oa_url"]
    for loc in work.get("locations", []):
        if loc.get("pdf_url"):
            return loc["pdf_url"]
    return None


def download_pdf(url: str, doi: str) -> Optional[Path]:
    PDF_CACHE_DIR.mkdir(exist_ok=True)
    safe = (doi.replace("/", "_").replace(":", "_")[:80] if doi else url[-40:])
    dest = PDF_CACHE_DIR / f"{safe}.pdf"
    if dest.exists():
        return dest
    try:
        headers = {"User-Agent": f"ECPC-Pipeline/1.0 (mailto:{OPENALEX_EMAIL})"}
        resp = requests.get(url, headers=headers, timeout=30, allow_redirects=True)
        if resp.status_code == 200 and b"%PDF" in resp.content[:1024]:
            dest.write_bytes(resp.content)
            return dest
    except Exception as e:
        print(f"    [WARN] PDF download failed: {e}")
    return None


def fetch_works(author_id: str, orcid: Optional[str] = None) -> list[dict]:
    """Fetch publications from OpenAlex for one author in the configured date range."""
    date_filter = f"from_publication_date:{DATE_FROM},to_publication_date:{DATE_TO}"
    if orcid:
        orcid_clean = orcid.replace("https://orcid.org/", "").strip()
        filter_str = f"author.orcid:{orcid_clean},{date_filter}"
    else:
        filter_str = f"author.id:{author_id},{date_filter}"

    url = "https://api.openalex.org/works"
    params = {
        "filter": filter_str,
        "mailto": OPENALEX_EMAIL,
        "per-page": 200,
        "select": (
            "id,doi,title,publication_year,publication_date,type,"
            "primary_location,locations,open_access,abstract_inverted_index,"
            "authorships,primary_topic,keywords,cited_by_count"
        ),
        "cursor": "*",
    }

    works: list[dict] = []
    while True:
        try:
            resp = requests.get(url, params=params, timeout=20)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            print(f"  [WARN] Works fetch failed: {e}")
            break

        batch = data.get("results", [])
        works.extend(batch)
        cursor = data.get("meta", {}).get("next_cursor")
        if not cursor or not batch:
            break
        params["cursor"] = cursor

    return works


# ── Claude extraction ──────────────────────────────────────────────────────────

EXTRACTION_SCHEMA = {
    "country": "Country/countries where the study was conducted",
    "setting_location": "Setting (e.g., school, community, refugee camp, clinic)",
    "latitude": "Approximate latitude of study location (float or null)",
    "longitude": "Approximate longitude of study location (float or null)",
    "longitudinal": "Longitudinal study? (yes / no / unclear)",
    "design": "Study design (RCT, quasi-experimental, cross-sectional, longitudinal cohort, qualitative, mixed-methods, systematic review, meta-analysis, other)",
    "followup_months": "Follow-up duration in months (integer or null)",
    "quant_qual": "Quantitative / qualitative / mixed",
    "type": "Study type (intervention / observational / review / other)",
    "intervention_name": "Name of intervention if applicable (or null)",
    "control": "Control or comparison condition (or null)",
    "randomized": "Randomization used? (yes / no / not applicable)",
    "sample_size": "Total sample size (integer or null)",
    "age_0_8": "N or % of children aged 0-8 in sample (string or null)",
    "age_9_25": "N or % aged 9-25 (string or null)",
    "age_adult": "N or % of adults (string or null)",
    "age_caregiver": "N or % who are caregivers (string or null)",
    "age_teacher": "N or % who are teachers/educators (string or null)",
    "demographics": "Key demographics: ethnicity, SES, displacement status, etc.",
    "child_reading": "Reading/literacy outcomes measured? (yes / no)",
    "child_writing": "Writing outcomes measured? (yes / no)",
    "child_math": "Math/numeracy outcomes measured? (yes / no)",
    "child_cognitive": "Cognitive development outcomes measured? (yes / no)",
    "child_language": "Language development outcomes measured? (yes / no)",
    "child_executive_function": "Executive function outcomes measured? (yes / no)",
    "child_motor": "Motor development outcomes measured? (yes / no)",
    "child_social_emotional": "Social-emotional outcomes measured? (yes / no)",
    "child_mental_health": "Mental health outcomes measured? (yes / no)",
    "child_behavioral": "Behavioral outcomes measured? (yes / no)",
    "child_physical": "Physical health outcomes measured? (yes / no)",
    "child_victimization": "Victimization/violence exposure outcomes measured? (yes / no)",
    "caregiver_outcomes": "Caregiver/family outcomes measured (brief description or none)",
    "teacher_outcomes": "Teacher/classroom outcomes measured (brief description or none)",
    "biomarkers": "Biomarker measures used (brief description or none)",
    "findings_psychometrics": "Psychometric/validation findings? (yes / no)",
    "findings_descriptive": "Descriptive/prevalence findings? (yes / no)",
    "findings_group_differences": "Group differences findings? (yes / no)",
    "findings_associations": "Association/correlation findings? (yes / no)",
    "findings_moderation": "Moderation findings? (yes / no)",
    "findings_mediation": "Mediation findings? (yes / no)",
    "findings_intervention_impact": "Intervention impact/effect findings? (yes / no)",
    "key_finding_1": "First key finding (1-2 sentences)",
    "key_finding_2": "Second key finding (1-2 sentences, or null)",
    "key_finding_3": "Third key finding (1-2 sentences, or null)",
    "effect_sizes": "Notable effect sizes or statistical results (or null)",
    "notes": "Methodological notes, caveats, or limitations",
}


def extract_with_claude(
    client: anthropic.Anthropic,
    work: dict,
    pdf_path: Optional[Path],
    abstract: str,
) -> dict:
    """Call Claude to extract structured metadata from a paper."""
    authors_str = ", ".join(
        a["author"]["display_name"] for a in work.get("authorships", [])[:10]
    )
    venue = ""
    pl = work.get("primary_location") or {}
    if pl.get("source"):
        venue = pl["source"].get("display_name", "")

    schema_json = json.dumps(EXTRACTION_SCHEMA, indent=2)

    system_prompt = (
        "You are a systematic review assistant extracting structured metadata "
        "for the ECPC (Early Childhood Peacebuilding Colloquium) Research Repository. "
        "The repository focuses on early childhood development in conflict-affected settings, "
        "peacebuilding interventions, and related topics. "
        "Always return ONLY a valid JSON object — no markdown fences, no explanations."
    )

    user_text = f"""Extract metadata from the paper below and return a JSON object with EXACTLY these keys:

{schema_json}

Paper details:
- Title: {work.get('title', 'Unknown')}
- Authors: {authors_str or 'Unknown'}
- Year: {work.get('publication_year', 'N/A')}
- Journal: {venue or 'N/A'}
- DOI: {work.get('doi', 'N/A')}
- Abstract: {abstract[:3000] if abstract else 'Not available'}

Rules:
- Use null (not the string "null") for missing numeric fields.
- Use "unclear" for text fields where the paper does not provide enough information.
- For yes/no fields, always answer exactly "yes" or "no" (lowercase).
- Do not add keys that are not in the schema."""

    content: list[dict] = []

    if pdf_path and pdf_path.exists():
        pdf_b64 = base64.standard_b64encode(pdf_path.read_bytes()).decode("utf-8")
        content.append(
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": pdf_b64,
                },
            }
        )
        user_text = user_text.replace(
            "Extract metadata from the paper below",
            "Extract metadata from the full PDF above (and supplementary details below)",
        )

    content.append({"type": "text", "text": user_text})

    response = client.messages.create(
        model=MODEL,
        max_tokens=2048,
        thinking={"type": "adaptive"},
        system=system_prompt,
        messages=[{"role": "user", "content": content}],
    )

    raw = ""
    for block in response.content:
        if block.type == "text":
            raw = block.text
            break

    # Parse JSON — strip any accidental markdown fences
    text = raw.strip()
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
    text = text.strip().rstrip("`").strip()

    try:
        start, end = text.find("{"), text.rfind("}") + 1
        if start >= 0 and end > start:
            return json.loads(text[start:end])
    except json.JSONDecodeError as e:
        print(f"    [WARN] JSON parse error: {e}")

    return {"_parse_error": raw[:300]}


# ── Keyword-based extraction (no API key required) ─────────────────────────────


def keyword_extract(work: dict, abstract: str) -> dict:
    """Rule-based metadata extraction from title + abstract text."""
    title = (work.get("title") or "").lower()
    text = title + " " + abstract.lower() if abstract else title
    work_type = (work.get("type") or "").lower()

    def has(*terms: str) -> bool:
        return any(t in text for t in terms)

    def yn(*terms: str) -> str:
        return "yes" if has(*terms) else "no"

    # ── Design
    if has("randomized controlled", "randomised controlled", " rct ", "(rct)", "randomly assigned", "random assignment", "random allocation"):
        design, randomized = "RCT", "yes"
    elif has("quasi-experimental", "quasi experimental", "difference-in-difference", "regression discontinuity"):
        design, randomized = "quasi-experimental", "no"
    elif has("meta-analysis", "meta analysis"):
        design, randomized = "meta-analysis", "not applicable"
    elif has("systematic review", "scoping review"):
        design, randomized = "systematic review", "not applicable"
    elif has("longitudinal", "prospective cohort", "cohort study"):
        design, randomized = "longitudinal cohort", "no"
    elif has("cross-sectional", "cross sectional"):
        design, randomized = "cross-sectional", "no"
    elif has("qualitative", "ethnograph", "grounded theory", "thematic analysis"):
        design, randomized = "qualitative", "not applicable"
    elif has("mixed method", "mixed-method"):
        design, randomized = "mixed-methods", "no"
    elif work_type in ("review",):
        design, randomized = "review", "not applicable"
    else:
        design, randomized = "unclear", "unclear"

    # ── Quant / qual
    is_qual = has("qualitative", "thematic analysis", "grounded theory", "ethnograph", "interview", "focus group", "phenomenolog")
    is_quant = has("quantitative", "regression", "survey", "scale", "score", "sample size", "n =", "n=")
    if is_qual and is_quant:
        quant_qual = "mixed"
    elif is_qual:
        quant_qual = "qualitative"
    else:
        quant_qual = "quantitative"

    # ── Longitudinal
    longitudinal = "yes" if has("longitudinal", "follow-up", "follow up", "over time", "prospective", "repeated measure", "waves of data") else "no"

    # ── Study type
    if has("systematic review", "meta-analysis", "scoping review", "literature review"):
        study_type = "review"
    elif has("intervention", "program", "programme", "training", "treatment", "therapy", "curriculum"):
        study_type = "intervention"
    else:
        study_type = "observational"

    # ── Child outcome domains
    child_reading = yn("reading", "literacy", "phonics", "decod", "reading skill", "print awareness")
    child_writing = yn("writing", "written language", "writing skill", "composition")
    child_math = yn("math", "numeracy", "arithmetic", "number skill", "mathematical")
    child_cognitive = yn("cognitive", "cognition", "intellectual", "intelligence", " iq ", "problem-solving", "problem solving")
    child_language = yn("language", "vocabulary", "communication", "speech", "linguistic", "bilingual")
    child_executive_function = yn("executive function", "self-regulation", "inhibitory control", "working memory", "effortful control", "cognitive control")
    child_motor = yn("motor", "physical development", "gross motor", "fine motor", "movement skill")
    child_social_emotional = yn("social-emotional", "social emotional", "social skill", "emotional", "empathy", "prosocial", " sel ", "social competence")
    child_mental_health = yn("mental health", "depression", "anxiety", "ptsd", "post-traumatic", "trauma", "wellbeing", "well-being", "psychological distress", "internalizing", "internalising")
    child_behavioral = yn("behavioral", "behavioural", "conduct", "aggression", "externalizing", "externalising", "disruptive", "attention deficit", "adhd")
    child_physical = yn("physical health", "nutrition", "stunting", "wasting", "growth", "weight-for-height", "height-for-age", "anthropometric")
    child_victimization = yn("violence", "abuse", "victimization", "victimisation", "maltreatment", "neglect", "corporal punishment", "physical punishment")

    # ── Other outcomes (brief descriptions)
    caregiver_parts: list[str] = []
    if has("caregiver", "parent", "mother", "father", "parenting"):
        if has("mental health", "depression", "stress", "wellbeing", "well-being", "distress"):
            caregiver_parts.append("mental health/wellbeing")
        if has("parenting practice", "parenting behavior", "parenting behaviour", "parenting skill"):
            caregiver_parts.append("parenting practices")
        if has("sensitivity", "responsiveness", "warmth", "stimulation"):
            caregiver_parts.append("sensitivity/responsiveness")
        if has("employment", "economic", "income"):
            caregiver_parts.append("economic outcomes")
        if not caregiver_parts:
            caregiver_parts.append("yes (see abstract)")
    caregiver_outcomes = "; ".join(caregiver_parts)

    teacher_outcomes = "yes (see abstract)" if has("teacher", "educator", "classroom", "teaching practice", "pedagogical") else ""

    biomarker_parts: list[str] = []
    if has("cortisol"): biomarker_parts.append("cortisol")
    if has(" eeg ", "electroencephalog"): biomarker_parts.append("EEG")
    if has("fmri", "neuroimaging", "brain imaging"): biomarker_parts.append("neuroimaging")
    if has("dna methylation", "epigenetic"): biomarker_parts.append("epigenetics")
    if has("telomere"): biomarker_parts.append("telomeres")
    if has("biomarker") and not biomarker_parts: biomarker_parts.append("biomarker (see abstract)")
    biomarkers = "; ".join(biomarker_parts)

    # ── Findings type
    findings_psychometrics = yn("validity", "reliability", "psychometric", "factor analysis", "validation", "measurement invariance", "internal consistency")
    findings_descriptive = yn("prevalence", "descriptive", "characteristics", "proportion", "percentage", "distribution")
    findings_group_differences = yn("group difference", "between group", "compared to", "significantly higher", "significantly lower", "t-test", "anova", "group comparison")
    findings_associations = yn("association", "correlation", "predict", "associated with", "related to", "regression", "odds ratio", "hazard ratio")
    findings_moderation = yn("moderation", "moderator", "moderated by", "interaction effect", "moderated mediation")
    findings_mediation = yn("mediation", "mediator", "mediated by", "indirect effect", "mediation analysis")
    findings_intervention_impact = yn("intervention effect", "program effect", "treatment effect", "significantly improved", "significant improvement", "efficacy", "effectiveness", "impact of the")

    return {
        "design": design,
        "randomized": randomized,
        "quant_qual": quant_qual,
        "longitudinal": longitudinal,
        "type": study_type,
        "child_reading": child_reading,
        "child_writing": child_writing,
        "child_math": child_math,
        "child_cognitive": child_cognitive,
        "child_language": child_language,
        "child_executive_function": child_executive_function,
        "child_motor": child_motor,
        "child_social_emotional": child_social_emotional,
        "child_mental_health": child_mental_health,
        "child_behavioral": child_behavioral,
        "child_physical": child_physical,
        "child_victimization": child_victimization,
        "caregiver_outcomes": caregiver_outcomes,
        "teacher_outcomes": teacher_outcomes,
        "biomarkers": biomarkers,
        "findings_psychometrics": findings_psychometrics,
        "findings_descriptive": findings_descriptive,
        "findings_group_differences": findings_group_differences,
        "findings_associations": findings_associations,
        "findings_moderation": findings_moderation,
        "findings_mediation": findings_mediation,
        "findings_intervention_impact": findings_intervention_impact,
    }


# ── Excel output ───────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    # Identity
    "ecpc_last", "ecpc_first", "ecpc_group", "ecpc_affiliation",
    # Citation
    "title", "authors", "year", "journal", "doi", "oa_url", "openalex_id", "cited_by_count",
    # Location
    "country", "setting_location", "latitude", "longitude",
    # Design
    "longitudinal", "design", "followup_months", "quant_qual", "type",
    "intervention_name", "control", "randomized", "sample_size",
    # Sample
    "age_0_8", "age_9_25", "age_adult", "age_caregiver", "age_teacher", "demographics",
    # Child outcomes
    "child_reading", "child_writing", "child_math", "child_cognitive", "child_language",
    "child_executive_function", "child_motor", "child_social_emotional",
    "child_mental_health", "child_behavioral", "child_physical", "child_victimization",
    # Other outcomes
    "caregiver_outcomes", "teacher_outcomes", "biomarkers",
    # Findings type
    "findings_psychometrics", "findings_descriptive", "findings_group_differences",
    "findings_associations", "findings_moderation", "findings_mediation",
    "findings_intervention_impact",
    # Key findings
    "key_finding_1", "key_finding_2", "key_finding_3", "effect_sizes",
    # Misc
    "notes", "abstract",
]

_WIDE_COLS = {
    "title", "abstract", "key_finding_1", "key_finding_2", "key_finding_3",
    "intervention_name", "demographics", "notes", "caregiver_outcomes",
    "teacher_outcomes", "biomarkers", "effect_sizes", "oa_url",
}
_MED_COLS = {"country", "setting_location", "design", "authors", "affiliation"}


def write_to_excel(records: list[dict], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ECPC Repository"

    header_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    ws.row_dimensions[1].height = 32

    for row_idx, record in enumerate(records, 2):
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            val = record.get(col_name)
            if val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.alignment = wrap

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if col_name in _WIDE_COLS:
            ws.column_dimensions[letter].width = 48
        elif col_name in _MED_COLS:
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 16

    ws.freeze_panes = "E2"  # freeze ECPC identity columns
    wb.save(out_path)
    print(f"\nSaved {len(records)} records → {out_path}")


# ── Row builder ────────────────────────────────────────────────────────────────


def build_row(member: dict, work: dict, extracted: dict) -> dict:
    abstract = reconstruct_abstract(work.get("abstract_inverted_index"))
    authors = ", ".join(
        a["author"]["display_name"] for a in work.get("authorships", [])[:15]
    )
    venue = ""
    pl = work.get("primary_location") or {}
    if pl.get("source"):
        venue = pl["source"].get("display_name", "")

    row: dict = {
        "ecpc_last": member["last"],
        "ecpc_first": member["first"],
        "ecpc_group": member.get("ecpc_group", ""),
        "ecpc_affiliation": member.get("affiliation", ""),
        "openalex_id": work.get("id", ""),
        "doi": work.get("doi", ""),
        "oa_url": get_oa_pdf_url(work) or "",
        "cited_by_count": work.get("cited_by_count", ""),
        "abstract": abstract[:2000] if abstract else "",
    }

    row.update(extracted)

    # OpenAlex ground truth overrides Claude where available
    if work.get("title"):
        row["title"] = work["title"]
    if work.get("publication_year"):
        row["year"] = work["publication_year"]
    if authors:
        row["authors"] = authors
    if venue:
        row["journal"] = venue
    if work.get("doi"):
        row["doi"] = work["doi"]

    return row


# ── Main pipeline ──────────────────────────────────────────────────────────────


def run_pipeline(
    members: list[dict],
    dry_run: bool = False,
    filter_last: Optional[str] = None,
    output: str = "ecpc_repository.xlsx",
) -> None:
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    use_claude = bool(api_key) and not dry_run
    client = anthropic.Anthropic() if use_claude else None

    if dry_run:
        print("Mode: dry-run (paper list only, no extraction)")
    elif use_claude:
        print("Mode: Claude extraction (full metadata)")
    else:
        print("Mode: keyword extraction (no API key — rule-based metadata)")

    all_records: list[dict] = []

    for member in members:
        last = member["last"].strip()
        first = member["first"].strip()

        if filter_last and last.lower() != filter_last.lower():
            continue

        print(f"\n{'─'*60}")
        print(f"  {first} {last}  |  {member.get('ecpc_group', '')}  |  {member.get('affiliation', '')}")

        orcid = member.get("orcid", "").strip()
        author_id: Optional[str] = None

        if orcid:
            print(f"  Using ORCID: {orcid}")
        else:
            author_id = find_openalex_author_id(last, first, member.get("affiliation", ""))
            if not author_id:
                print(f"  [SKIP] No OpenAlex author ID found — add ORCID to orcid_ids.csv for better results")
                continue
            print(f"  OpenAlex ID: {author_id}")

        works = fetch_works(author_id or "", orcid or None)
        print(f"  {len(works)} works found ({DATE_FROM} – {DATE_TO})")

        for work in works:
            title = (work.get("title") or "No title")[:80]
            abstract = reconstruct_abstract(work.get("abstract_inverted_index"))

            if dry_run:
                year = work.get("publication_year", "?")
                oa = "OA" if work.get("open_access", {}).get("is_oa") else "  "
                print(f"    [{year}] [{oa}] {title}")
                continue

            print(f"  → {title[:70]}...")

            if use_claude:
                pdf_path: Optional[Path] = None
                oa_url = get_oa_pdf_url(work)
                if oa_url:
                    pdf_path = download_pdf(oa_url, work.get("doi", ""))
                    if pdf_path:
                        print(f"    PDF cached: {pdf_path.name}")
                try:
                    extracted = extract_with_claude(client, work, pdf_path, abstract)
                except Exception as e:
                    print(f"    [WARN] Claude extraction failed: {e}")
                    extracted = keyword_extract(work, abstract)
            else:
                extracted = keyword_extract(work, abstract)

            all_records.append(build_row(member, work, extracted))
            time.sleep(0.3)

    if dry_run:
        print(f"\nDry-run complete. Re-run without --dry-run to extract metadata.")
    elif all_records:
        write_to_excel(all_records, output)
    else:
        print("\nNo records found. Check member names or add ORCID IDs.")


# ── CLI ────────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description=f"ECPC Research Repository Pipeline ({DATE_FROM} – {DATE_TO})",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--members-csv", default="members.csv",
        help="Path to members CSV (default: members.csv)",
    )
    parser.add_argument(
        "--orcid-csv", default="orcid_ids.csv",
        help="Path to ORCID IDs CSV (default: orcid_ids.csv)",
    )
    parser.add_argument(
        "--member", metavar="LAST_NAME",
        help="Run for a single member by last name (e.g. --member Connolly)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Fetch paper list only — no Claude API calls, no PDF downloads",
    )
    parser.add_argument(
        "--output", default="ecpc_repository.xlsx",
        help="Output Excel file path (default: ecpc_repository.xlsx)",
    )
    args = parser.parse_args()

    members_path = Path(args.members_csv)
    if not members_path.exists():
        # Try relative to this script's directory
        members_path = Path(__file__).parent / args.members_csv
    if not members_path.exists():
        print(f"ERROR: Members CSV not found: {args.members_csv}")
        sys.exit(1)

    if not args.dry_run and not os.environ.get("ANTHROPIC_API_KEY"):
        print("No ANTHROPIC_API_KEY found — running with keyword-based extraction.")

    orcid_path = Path(args.orcid_csv)
    if not orcid_path.exists():
        orcid_path = Path(__file__).parent / args.orcid_csv

    members = load_members(str(members_path), str(orcid_path) if orcid_path.exists() else None)
    print(f"Loaded {len(members)} ECPC members.")

    run_pipeline(
        members,
        dry_run=args.dry_run,
        filter_last=args.member,
        output=args.output,
    )


if __name__ == "__main__":
    main()
